#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import os
import openpyxl

#Prepare the spreadsheets to copy from and paste too.

#File to be copied
wb = openpyxl.load_workbook("Sample1.xlsx") #Add file name
copySheet = wb["Area 02"] #Add Sheet name

# #File to be pasted into
template = openpyxl.load_workbook("Sample2.xlsx") #Add file name
pasteSheet = template["Area 02"] #Add Sheet name

def searchForWord(sheet, theWord):
    for i in range(1, 11,1):
        #Appends the row to a RowSelected list
        for j in range(1, 11,1):
            if theWord in str(sheet.cell(row = i, column = j).value).upper():
                return j
tagPaste= searchForWord(pasteSheet, "TAG")
tagCopy= searchForWord(copySheet, "TAG")

problemPaste= searchForWord(pasteSheet,"PROBLEM")
problemCopy= searchForWord(copySheet,"PROBLEM")

complainPaste= searchForWord(pasteSheet,"COMP")
complainCopy= searchForWord(copySheet,"COMP")

actionPaste= searchForWord(pasteSheet,"ACTION")
actionCopy= searchForWord(copySheet,"ACTION")

statusPaste= searchForWord(pasteSheet,"COMP")
statusCopy= searchForWord(copySheet,"COMP")

datePaste= searchForWord(pasteSheet,"DATE")
dateCopy= searchForWord(copySheet,"DATE")
# pasteList= {"Tag": tagPaste, "Problem": problemPaste, "Action" : actionPaste, "Status": statusPaste}
# copyList= {"Tag": tagCopy, "Problem": problemCopy, "Action" : actionCopy, "Status": statusCopy}

pasteList=[tagPaste, problemPaste, complainPaste, actionPaste, statusPaste, datePaste]
copyList= [tagCopy, problemCopy, complainCopy, actionCopy, statusCopy, dateCopy]

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(copyList, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(12,14,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in copyList:
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
#print(copyRange(copyList, copySheet))
#Paste range

#Paste data from copyRange into template sheet
def pasteRange(pasteList, sheetReceiving,copiedData):
    countRow = 0
    for i in range(12,14,1):
        countCol = 0
        for j in pasteList:

            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def createData():
    print("Processing...")
    selectedRange = copyRange(copyList,copySheet) #Change the 4 number values
    pasteRange(pasteList, pasteSheet, selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("Sample2.xlsx")
    print("items copied and pasted!")

createData()
#open all the files in a folder
# x= input("enter the path:")
# print(x)
# os.chdir(x)
# f= os.listdir(x)
# for i in f:
#     if i.endswith('.py') or i.endswith('.xlsx'):
#         print(i)

